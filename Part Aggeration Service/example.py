import requests
import json
import pandas as pd
from typing import Optional, Dict, Any, List, Tuple
import time
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import logging
import chardet

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PASAPIClient:
    def __init__(self, config_file_path: Optional[str] = None, bearer_token: Optional[str] = None):
        """
        Initialize the PAS API client
        
        Args:
            config_file_path: Path to JSON config file (optional)
            bearer_token: Pre-obtained bearer token (optional)
        """
        # Default configuration
        self.config = {
            "PasUrl": "https://api.pas.partquest.com",
            "AuthServiceUrl": "https://samauth.us-east-1.sws.siemens.com/",
            "ClientId": "WQAVuBZ8aENg-LFNzbkL2",
            "ClientSecret": "EBEipGnyFVCSy0wrBH6mr3PJD2J63JMj4Tb2FVIHAVzo0OH8ao_uJpEOdo-aJBbcPbSs1jI00RfG18rMnpcyDQ",
            "SearchProviderId": "44",
            "SearchProviderVersion": "2",
            "MultiMatchResultsLimit": 10,
            "EnableRawOutput": True,
            "EnableEnriching": True,
            "Resilience": {
                "Retry": {
                    "RetryCount": 2,
                    "MedianFirstRetryDelay": "00:00:01.000"
                },
                "Timeout": "00:01:00.000"
            }
        }
        
        # Load from file if provided
        if config_file_path and os.path.exists(config_file_path):
            with open(config_file_path, 'r') as f:
                file_config = json.load(f)
                self.config.update(file_config)
        
        # If a bearer token is provided, use it directly
        if bearer_token:
            self.access_token = bearer_token
            # Set expiration to 2 hours from now (typical token lifetime)
            self.token_expires_at = datetime.now() + timedelta(hours=2)
        else:
            self.access_token = None
            self.token_expires_at = None
        
        # Supply Chain enricher configuration
        self.SUPPLY_CHAIN_ID = "33"
        self.SUPPLY_CHAIN_VERSION = "1"
        
        # Store available enrichers
        self.available_enrichers = {}
        
        # Store raw responses for debugging
        self.raw_responses = []
    
    def get_all_enrichers(self):
        """Get all available enriching providers"""
        endpoint = '/api/v2/enriching-providers/get-all'
        
        try:
            result = self._make_request('GET', endpoint)
            if result.get('success') and result.get('result'):
                enrichers = result['result'].get('enrichingProviders', [])
                logger.info(f"\nFound {len(enrichers)} enriching providers:")
                
                for provider in enrichers:
                    provider_id = provider['id']
                    provider_name = provider['name']
                    logger.info(f"  ID: {provider_id} - Name: {provider_name}")
                    
                    # Get versions for each enricher
                    versions = self.get_enricher_versions(provider_id)
                    if versions:
                        self.available_enrichers[provider_id] = {
                            'name': provider_name,
                            'versions': versions
                        }
                
                return enrichers
        except Exception as e:
            logger.error(f"Failed to get enriching providers: {e}")
        return []
    
    def get_enricher_versions(self, enricher_id: str):
        """Get available versions for an enricher"""
        endpoint = f'/api/v2/enriching-providers/{enricher_id}/get-versions'
        
        try:
            result = self._make_request('GET', endpoint)
            if result.get('success') and result.get('result'):
                versions = result['result'].get('versions', [])
                active_versions = [str(v['number']) for v in versions if v.get('isActive')]
                return active_versions
        except Exception as e:
            logger.debug(f"Failed to get versions for enricher {enricher_id}: {e}")
        return []
    
    def get_enricher_properties(self, enricher_id: str, version: str):
        """Get properties for a specific enricher"""
        endpoint = f'/api/v2/enriching-providers/{enricher_id}/{version}/get-definition'
        
        try:
            result = self._make_request('GET', endpoint)
            if result.get('success') and result.get('result'):
                properties = result['result'].get('propertyDefinitions', [])
                
                # Look for price-related properties
                price_properties = []
                for prop in properties:
                    prop_name = prop['name'].lower()
                    if any(word in prop_name for word in ['price', 'cost', 'pricing', 'msrp', 'list']):
                        price_properties.append(prop)
                        logger.info(f"    Found price property: {prop['id']} - {prop['name']}")
                
                return properties, price_properties
        except Exception as e:
            logger.debug(f"Failed to get properties for enricher {enricher_id}/{version}: {e}")
        return [], []
    
    def scan_for_pricing_enrichers(self):
        """Scan all enrichers to find ones with pricing data"""
        logger.info("\n" + "="*60)
        logger.info("Scanning for enrichers with pricing data...")
        logger.info("="*60)
        
        # Get all enrichers
        self.get_all_enrichers()
        
        pricing_enrichers = {}
        
        for enricher_id, enricher_info in self.available_enrichers.items():
            logger.info(f"\nChecking enricher: {enricher_info['name']} (ID: {enricher_id})")
            
            for version in enricher_info['versions']:
                logger.info(f"  Version {version}:")
                all_props, price_props = self.get_enricher_properties(enricher_id, version)
                
                if price_props:
                    pricing_enrichers[enricher_id] = {
                        'name': enricher_info['name'],
                        'version': version,
                        'price_properties': price_props,
                        'all_properties': all_props
                    }
        
        if pricing_enrichers:
            logger.info("\n" + "="*60)
            logger.info("FOUND ENRICHERS WITH PRICING DATA:")
            for eid, info in pricing_enrichers.items():
                logger.info(f"  {info['name']} (ID: {eid}, Version: {info['version']})")
                for prop in info['price_properties']:
                    logger.info(f"    - {prop['id']}: {prop['name']}")
        else:
            logger.info("\nNo enrichers with obvious pricing properties found.")
            logger.info("Pricing may need to be obtained directly from distributors.")
        
        return pricing_enrichers
    
    def _detect_encoding(self, file_path: str) -> str:
        """Detect the encoding of a file"""
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)  # Read first 10KB
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            confidence = result['confidence']
            logger.info(f"Detected encoding: {encoding} (confidence: {confidence:.2%})")
            return encoding if encoding else 'utf-8'
    
    def _get_access_token(self) -> str:
        """Get or refresh the access token"""
        if self.access_token and self.token_expires_at:
            if datetime.now() < self.token_expires_at:
                return self.access_token
        
        logger.info("Obtaining new access token...")
        
        # Use the correct token endpoint
        token_url = f"{self.config['AuthServiceUrl']}token"
        
        # Use basic auth with client credentials
        auth = (self.config['ClientId'], self.config['ClientSecret'])
        
        auth_data = {
            'grant_type': 'client_credentials',
            'scope': 'sws.icarus.api.read'  # Required scope for PAS API
        }
        
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        
        try:
            response = requests.post(
                token_url, 
                auth=auth,
                data=auth_data,
                headers=headers,
                timeout=10
            )
            response.raise_for_status()
            
            token_data = response.json()
            self.access_token = token_data['access_token']
            
            expires_in = token_data.get('expires_in', 7200)  # Default to 2 hours
            self.token_expires_at = datetime.now() + timedelta(seconds=expires_in - 60)
            
            logger.info("Access token obtained successfully")
            logger.info(f"Token expires in {expires_in} seconds")
            return self.access_token
            
        except Exception as e:
            logger.error(f"Authentication failed: {e}")
            raise
    
    def _make_request(self, method: str, endpoint: str, 
                     json_data: Optional[Dict] = None,
                     params: Optional[Dict] = None) -> Dict[str, Any]:
        """Make an authenticated request with retry logic"""
        
        retry_count = self.config['Resilience']['Retry']['RetryCount']
        for attempt in range(retry_count + 1):
            try:
                token = self._get_access_token()
                
                # Include all required headers
                headers = {
                    'Authorization': f'Bearer {token}',
                    'Content-Type': 'application/json',
                    'X-Siemens-Correlation-Id': f'corr-{int(time.time() * 1000)}',
                    'X-Siemens-Session-Id': f'session-{int(time.time())}',
                    'X-Siemens-Ebs-User-Country-Code': 'US',
                    'X-Siemens-Ebs-User-Currency': 'USD'
                }
                
                url = f"{self.config['PasUrl']}{endpoint}"
                
                logger.debug(f"Making {method} request to {url}")
                if json_data:
                    logger.debug(f"Request body: {json.dumps(json_data, indent=2)}")
                
                response = requests.request(
                    method=method,
                    url=url,
                    headers=headers,
                    json=json_data,
                    params=params,
                    timeout=60
                )
                
                if response.status_code == 401:
                    logger.warning("Token expired, getting new token...")
                    self.access_token = None
                    self.token_expires_at = None
                    continue
                
                response.raise_for_status()
                result = response.json()
                
                if not result.get('success', False):
                    error = result.get('error', {})
                    error_msg = error.get('message', 'Unknown error')
                    
                    # Log detailed error information
                    logger.error(f"API returned success=false: {error_msg}")
                    if 'causes' in error and error['causes']:
                        logger.error("Error details:")
                        for cause in error['causes']:
                            cause_code = cause.get('code', 'Unknown')
                            cause_msg = cause.get('message', 'No message')
                            logger.error(f"  - {cause_code}: {cause_msg}")
                    
                    raise Exception(f"API Error: {error_msg}")
                
                return result
                
            except Exception as e:
                if attempt < retry_count:
                    logger.warning(f"Request failed (attempt {attempt + 1}/{retry_count + 1}): {e}")
                    time.sleep(1 * (attempt + 1))
                else:
                    raise
        
        raise Exception("Request failed after all retry attempts")
    
    def extract_complex_value(self, value):
        """Extract value from complex objects including pricing data"""
        if isinstance(value, dict):
            if value.get('__complex__') == 'Url':
                return value.get('value', 'N/A')
            elif value.get('__complex__') == 'Timestamp':
                return value.get('iso8601Timestamp', 'N/A')
            elif value.get('__complex__') == 'CustomV2':
                # Extract pricing information
                fields = value.get('fields', {}).get('succeeded', {})
                if 'Value' in fields and 'Quantity' in fields:
                    price = fields.get('Value', 'N/A')
                    qty = fields.get('Quantity', 'N/A')
                    currency = fields.get('Currency', 'USD')
                    return {'price': price, 'quantity': qty, 'currency': currency}
                return str(fields) if fields else 'N/A'
            elif 'value' in value:
                return str(value.get('value', 'N/A'))
            else:
                return str(value)
        elif isinstance(value, list):
            # Handle list of pricing tiers
            if value and isinstance(value[0], dict) and value[0].get('__complex__') == 'CustomV2':
                pricing_tiers = []
                for item in value:
                    extracted = self.extract_complex_value(item)
                    if isinstance(extracted, dict) and 'price' in extracted:
                        pricing_tiers.append(extracted)
                return pricing_tiers
            return value
        return str(value) if value else 'N/A'
    
    def extract_enriching_data_improved(self, part_data: Dict) -> Dict:
        """Extract all enriching data including detailed pricing"""
        enriched_data = {
            'distributors': []
        }
        
        if 'enrichingResult' in part_data and part_data['enrichingResult']:
            enriching_result = part_data['enrichingResult']
            
            if 'succeeded' in enriching_result and '33' in enriching_result['succeeded']:
                supply_chain_data = enriching_result['succeeded']['33']
                
                if 'matchedRecords' in supply_chain_data:
                    for record in supply_chain_data['matchedRecords']:
                        if 'properties' in record and 'succeeded' in record['properties']:
                            props = record['properties']['succeeded']
                            
                            distributor_info = {
                                'name': self.extract_complex_value(props.get('339c3014', '')),
                                'authorized': props.get('08c82fa6', False),
                                'stock': self.extract_complex_value(props.get('3708193e', 'N/A')),
                                'lead_time': self.extract_complex_value(props.get('198a2ad2', 'N/A')),
                                'availability': self.extract_complex_value(props.get('482adab0', 'N/A')),
                                'description': self.extract_complex_value(props.get('bf4dd752', 'N/A')),
                                'datasheet_url': self.extract_complex_value(props.get('24538207', 'N/A')),
                                'buy_url': self.extract_complex_value(props.get('db80a0d0', 'N/A')),
                                'pb_free': self.extract_complex_value(props.get('e677f6f7', 'N/A')),
                                'last_updated': self.extract_complex_value(props.get('fdd91810', 'N/A')),
                                'distributor_pn': self.extract_complex_value(props.get('8f6be867', 'N/A')),
                                'risk_rank': self.extract_complex_value(props.get('ae390795', 'N/A')),
                                'pricing': [],
                                'estimated_pricing': []
                            }
                            
                            # Extract actual pricing (5702a948)
                            if '5702a948' in props:
                                pricing = self.extract_complex_value(props['5702a948'])
                                if isinstance(pricing, list):
                                    distributor_info['pricing'] = pricing
                            
                            # Extract estimated pricing (a0ecfd70)
                            if 'a0ecfd70' in props:
                                est_pricing = self.extract_complex_value(props['a0ecfd70'])
                                if isinstance(est_pricing, list):
                                    distributor_info['estimated_pricing'] = est_pricing
                            
                            # Only add if we have a valid distributor name
                            if distributor_info['name'] and distributor_info['name'] != 'N/A':
                                enriched_data['distributors'].append(distributor_info)
        
        return enriched_data
    
    def search_part_with_enriching(self, manufacturer_pn: str, manufacturer: str, pricing_enrichers: Dict = None) -> Tuple[Dict, str]:
        """
        Search for a part with enriching data
        """
        search_term = manufacturer_pn
        
        logger.info(f"Searching with enriching for: {manufacturer} {manufacturer_pn}")
        
        endpoint = f'/api/v2/search-providers/{self.config["SearchProviderId"]}/{self.config["SearchProviderVersion"]}/free-text/search'
        
        # Supply Chain enricher properties
        supply_chain_outputs = [
            "951ed6a7",  # Distributor Id
            "339c3014",  # Distributor Name
            "08c82fa6",  # Distributor Authorized Status
            "482adab0",  # Availability
            "198a2ad2",  # Lead Time
            "ae390795",  # Risk Rank
            "3708193e",  # Stock Indicator
            "bf4dd752",  # Description
            "8f6be867",  # Distributor Part Number
            "24538207",  # Datasheet Url
            "db80a0d0",  # Buy Now Url
            "e677f6f7",  # Pb Free
            "fdd91810",  # Last Updated
        ]
        
        request_body = {
            "ftsParameters": {
                "match": {
                    "term": search_term
                },
                "paging": {
                    "requestedPageSize": 20
                }
            }
        }
        
        # Add enriching parameters if enabled
        if self.config.get('EnableEnriching', True):
            enriching_params = []
            
            # Check if Supply Chain enricher (ID 33) has pricing properties
            if pricing_enrichers and self.SUPPLY_CHAIN_ID in pricing_enrichers:
                # Combine standard outputs with pricing properties for Supply Chain enricher
                combined_outputs = supply_chain_outputs.copy()
                
                # Add pricing property IDs
                price_prop_ids = [p['id'] for p in pricing_enrichers[self.SUPPLY_CHAIN_ID]['price_properties']]
                for prop_id in price_prop_ids:
                    if prop_id not in combined_outputs:
                        combined_outputs.append(prop_id)
                
                enriching_params.append({
                    "enrichingProviderId": self.SUPPLY_CHAIN_ID,
                    "enrichingProviderVersion": self.SUPPLY_CHAIN_VERSION,
                    "outputs": combined_outputs
                })
                
                logger.debug(f"Added Supply Chain enricher with combined outputs: {len(combined_outputs)} properties")
            else:
                # Just add standard Supply Chain enricher
                enriching_params.append({
                    "enrichingProviderId": self.SUPPLY_CHAIN_ID,
                    "enrichingProviderVersion": self.SUPPLY_CHAIN_VERSION,
                    "outputs": supply_chain_outputs
                })
            
            # Add any OTHER pricing enrichers (not ID 33)
            if pricing_enrichers:
                for enricher_id, enricher_info in pricing_enrichers.items():
                    if enricher_id != self.SUPPLY_CHAIN_ID:  # Skip Supply Chain enricher as we already handled it
                        # Get all property IDs from price properties
                        price_prop_ids = [p['id'] for p in enricher_info['price_properties']]
                        
                        enriching_params.append({
                            "enrichingProviderId": enricher_id,
                            "enrichingProviderVersion": enricher_info['version'],
                            "outputs": price_prop_ids
                        })
                        logger.debug(f"Added pricing enricher {enricher_id} with properties: {price_prop_ids}")
            
            request_body["enrichingParameters"] = enriching_params
        
        try:
            result = self._make_request('POST', endpoint, json_data=request_body)
            
            if result.get('result') and result['result'].get('results'):
                parts = result['result']['results']
                total_count = result['result'].get('totalCount', len(parts))
                
                logger.info(f"Found {total_count} total results")
                
                # Filter and find best matches
                exact_matches = []
                partial_matches = []
                
                for part_data in parts:
                    part = part_data.get('searchProviderPart', {})
                    found_mpn = part.get('manufacturerPartNumber', '')
                    found_mfr = part.get('manufacturerName', '')
                    
                    if found_mpn.upper() == manufacturer_pn.upper():
                        if manufacturer.upper() in found_mfr.upper() or found_mfr.upper() in manufacturer.upper():
                            exact_matches.append(part_data)
                        else:
                            partial_matches.append(part_data)
                
                if exact_matches:
                    return {
                        'result': {
                            'results': exact_matches[:self.config["MultiMatchResultsLimit"]], 
                            'totalCount': len(exact_matches)
                        }
                    }, 'exact'
                elif partial_matches:
                    return {
                        'result': {
                            'results': partial_matches[:self.config["MultiMatchResultsLimit"]], 
                            'totalCount': len(partial_matches)
                        }
                    }, 'partial'
                else:
                    return result, 'partial'
            else:
                return result, 'no_match'
                
        except Exception as e:
            logger.error(f"Search failed: {e}")
            return {'error': str(e)}, 'error'
    
    def generate_html_report(self, results: List[Dict], output_file: str):
        """Generate a scalable Siemens-branded HTML report for large datasets"""
        
        # Calculate statistics
        total_parts = len(results)
        exact_matches = sum(1 for r in results if r.get('match_type') == 'exact')
        partial_matches = sum(1 for r in results if r.get('match_type') == 'partial')
        no_matches = sum(1 for r in results if r.get('match_type') in ['no_match', 'error'])
        total_distributors = sum(len(r.get('distributors', [])) for r in results)
        
        # Find best prices across all parts
        best_prices = []
        for result in results:
            if result.get('distributors'):
                for dist in result['distributors']:
                    pricing = dist.get('pricing', []) or dist.get('estimated_pricing', [])
                    if pricing:
                        for tier in pricing:
                            if isinstance(tier, dict) and 'price' in tier:
                                best_prices.append({
                                    'part': f"{result.get('input_manufacturer', '')} {result.get('input_mpn', '')}",
                                    'distributor': dist.get('name', ''),
                                    'price': tier.get('price', 0),
                                    'quantity': tier.get('quantity', 0)
                                })
        
        # Sort best prices
        best_prices.sort(key=lambda x: x['price'])
        
        html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Siemens PAS API Part Search Report</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            
            :root {
                /* Siemens Brand Colors */
                --siemens-petrol-dark: #0D1E2B;
                --siemens-petrol: #003C4C;
                --siemens-teal: #2B9999;
                --siemens-teal-light: #00CFC1;
                --siemens-gray-light: #E5E5E5;
                --siemens-gray: #8B8B8B;
                --siemens-white: #FFFFFF;
                --success-green: #00B388;
                --warning-yellow: #FFB700;
                --error-red: #DB3333;
            }
            
            body {
                font-family: 'Siemens Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
                background: linear-gradient(135deg, var(--siemens-petrol-dark) 0%, var(--siemens-petrol) 100%);
                min-height: 100vh;
                color: #333;
            }
            
            .header {
                background: var(--siemens-petrol-dark);
                color: var(--siemens-white);
                padding: 2rem;
                box-shadow: 0 2px 10px rgba(0,0,0,0.2);
            }
            
            .header h1 {
                font-size: 2rem;
                font-weight: 300;
                letter-spacing: -0.5px;
                display: flex;
                align-items: center;
                gap: 1rem;
            }
            
            .siemens-logo {
                height: 40px;
            }
            
            .container {
                max-width: 1600px;
                margin: 0 auto;
                padding: 2rem;
            }
            
            .summary-section {
                background: var(--siemens-white);
                border-radius: 8px;
                padding: 2rem;
                margin-bottom: 2rem;
                box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            }
            
            .summary-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 1.5rem;
                margin-top: 1.5rem;
            }
            
            .stat-card {
                background: linear-gradient(135deg, var(--siemens-teal) 0%, var(--siemens-teal-light) 100%);
                color: var(--siemens-white);
                padding: 1.5rem;
                border-radius: 8px;
                text-align: center;
                transition: transform 0.3s;
            }
            
            .stat-card:hover {
                transform: translateY(-2px);
            }
            
            .stat-value {
                font-size: 2.5rem;
                font-weight: 300;
                margin-bottom: 0.5rem;
            }
            
            .stat-label {
                font-size: 0.9rem;
                opacity: 0.95;
            }
            
            .controls {
                background: var(--siemens-white);
                border-radius: 8px;
                padding: 1.5rem;
                margin-bottom: 2rem;
                box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            }
            
            .search-box {
                width: 100%;
                padding: 1rem;
                border: 2px solid var(--siemens-teal);
                border-radius: 4px;
                font-size: 1rem;
                margin-bottom: 1rem;
            }
            
            .filter-buttons {
                display: flex;
                gap: 0.5rem;
                flex-wrap: wrap;
            }
            
            .filter-btn {
                padding: 0.5rem 1.5rem;
                background: var(--siemens-gray-light);
                border: 2px solid transparent;
                border-radius: 4px;
                cursor: pointer;
                transition: all 0.3s;
                font-size: 0.9rem;
            }
            
            .filter-btn:hover {
                background: var(--siemens-teal-light);
                color: var(--siemens-white);
            }
            
            .filter-btn.active {
                background: var(--siemens-teal);
                color: var(--siemens-white);
                border-color: var(--siemens-teal);
            }
            
            .results-info {
                background: var(--siemens-white);
                padding: 1rem 1.5rem;
                border-radius: 8px;
                margin-bottom: 1rem;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }
            
            .pagination {
                display: flex;
                gap: 0.5rem;
                align-items: center;
            }
            
            .page-btn {
                padding: 0.5rem 1rem;
                background: var(--siemens-teal);
                color: var(--siemens-white);
                border: none;
                border-radius: 4px;
                cursor: pointer;
                transition: background 0.3s;
            }
            
            .page-btn:hover:not(:disabled) {
                background: var(--siemens-petrol);
            }
            
            .page-btn:disabled {
                background: var(--siemens-gray-light);
                cursor: not-allowed;
            }
            
            .parts-container {
                display: grid;
                gap: 1rem;
            }
            
            .part-card {
                background: var(--siemens-white);
                border-radius: 8px;
                overflow: hidden;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                transition: all 0.3s;
            }
            
            .part-card:hover {
                box-shadow: 0 4px 20px rgba(0,0,0,0.15);
            }
            
            .part-header {
                padding: 1rem 1.5rem;
                background: linear-gradient(90deg, var(--siemens-petrol-dark) 0%, var(--siemens-petrol) 100%);
                color: var(--siemens-white);
                display: flex;
                justify-content: space-between;
                align-items: center;
                cursor: pointer;
            }
            
            .part-title {
                font-size: 1.1rem;
                font-weight: 500;
            }
            
            .match-badge {
                padding: 0.25rem 0.75rem;
                border-radius: 20px;
                font-size: 0.8rem;
                font-weight: 500;
            }
            
            .match-exact {
                background: var(--success-green);
                color: var(--siemens-white);
            }
            
            .match-partial {
                background: var(--warning-yellow);
                color: var(--siemens-petrol-dark);
            }
            
            .match-no_match, .match-error {
                background: var(--error-red);
                color: var(--siemens-white);
            }
            
            .part-content {
                max-height: 0;
                overflow: hidden;
                transition: max-height 0.3s ease-out;
            }
            
            .part-card.expanded .part-content {
                max-height: 2000px;
                transition: max-height 0.5s ease-in;
            }
            
            .part-details {
                padding: 1.5rem;
                border-bottom: 1px solid var(--siemens-gray-light);
            }
            
            .distributors-section {
                padding: 1.5rem;
            }
            
            .distributors-grid {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                gap: 1rem;
                margin-top: 1rem;
            }
            
            .distributor-card {
                background: #f8f8f8;
                border: 1px solid var(--siemens-gray-light);
                border-radius: 6px;
                padding: 1rem;
            }
            
            .distributor-name {
                font-weight: 500;
                color: var(--siemens-petrol);
                margin-bottom: 0.5rem;
            }
            
            .price-table {
                width: 100%;
                margin-top: 0.5rem;
                font-size: 0.85rem;
            }
            
            .price-table th {
                background: var(--siemens-teal);
                color: var(--siemens-white);
                padding: 0.25rem 0.5rem;
                text-align: left;
            }
            
            .price-table td {
                padding: 0.25rem 0.5rem;
                border-bottom: 1px solid var(--siemens-gray-light);
            }
            
            .expand-icon {
                transition: transform 0.3s;
                display: inline-block;
            }
            
            .part-card.expanded .expand-icon {
                transform: rotate(180deg);
            }
            
            .loading {
                display: none;
                text-align: center;
                padding: 2rem;
                color: var(--siemens-white);
            }
            
            .export-btn {
                position: fixed;
                bottom: 2rem;
                right: 2rem;
                padding: 1rem 2rem;
                background: var(--siemens-teal);
                color: var(--siemens-white);
                border: none;
                border-radius: 50px;
                cursor: pointer;
                box-shadow: 0 4px 20px rgba(0,0,0,0.2);
                transition: all 0.3s;
                font-size: 1rem;
            }
            
            .export-btn:hover {
                background: var(--siemens-petrol);
                transform: translateY(-2px);
            }
            
            @media (max-width: 768px) {
                .container {
                    padding: 1rem;
                }
                
                .distributors-grid {
                    grid-template-columns: 1fr;
                }
            }
        </style>
    </head>
    <body>
        <div class="header">
            <div class="container">
                <h1>
                    <svg class="siemens-logo" viewBox="0 0 100 20" fill="currentColor">
                        <text x="0" y="15" font-family="Arial, sans-serif" font-size="16" font-weight="bold">SIEMENS</text>
                    </svg>
                    PAS API Part Search Report
                </h1>
            </div>
        </div>
        
        <div class="container">
            <div class="summary-section">
                <h2 style="color: var(--siemens-petrol); margin-bottom: 1rem;">Executive Summary</h2>
                <div class="summary-grid">
                    <div class="stat-card">
                        <div class="stat-value">""" + f"{total_parts:,}" + """</div>
                        <div class="stat-label">Total Parts Searched</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + f"{exact_matches:,}" + """</div>
                        <div class="stat-label">Exact Matches</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + f"{partial_matches:,}" + """</div>
                        <div class="stat-label">Partial Matches</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + f"{no_matches:,}" + """</div>
                        <div class="stat-label">No Matches</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + f"{total_distributors:,}" + """</div>
                        <div class="stat-label">Distributor Records</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">""" + f"{(exact_matches/total_parts*100):.1f}%" + """</div>
                        <div class="stat-label">Match Rate</div>
                    </div>
                </div>
            </div>
            
            <div class="controls">
                <input type="text" class="search-box" id="searchBox" placeholder="Search by part number, manufacturer, or distributor...">
                <div class="filter-buttons">
                    <button class="filter-btn active" data-filter="all">All Parts (""" + f"{total_parts:,}" + """)</button>
                    <button class="filter-btn" data-filter="exact">Exact Matches (""" + f"{exact_matches:,}" + """)</button>
                    <button class="filter-btn" data-filter="partial">Partial Matches (""" + f"{partial_matches:,}" + """)</button>
                    <button class="filter-btn" data-filter="no_match">No Matches (""" + f"{no_matches:,}" + """)</button>
                </div>
            </div>
            
            <div class="results-info">
                <span>Showing <span id="visibleCount">0</span> of <span id="totalCount">0</span> parts</span>
                <div class="pagination">
                    <button class="page-btn" id="prevBtn" onclick="changePage(-1)">Previous</button>
                    <span>Page <span id="currentPage">1</span> of <span id="totalPages">1</span></span>
                    <button class="page-btn" id="nextBtn" onclick="changePage(1)">Next</button>
                </div>
            </div>
            
            <div class="loading" id="loading">Loading parts...</div>
            <div class="parts-container" id="partsContainer"></div>
        </div>
        
        <button class="export-btn" onclick="exportToCSV()">Export Results to CSV</button>
        
        <script>
            // Parts data
            const allParts = """ + json.dumps(results) + """;
            let filteredParts = [...allParts];
            let currentPage = 1;
            const partsPerPage = 50;
            
            // Initialize
            document.addEventListener('DOMContentLoaded', function() {
                updateDisplay();
                setupEventListeners();
            });
            
            function setupEventListeners() {
                // Search functionality
                document.getElementById('searchBox').addEventListener('input', function(e) {
                    const searchTerm = e.target.value.toLowerCase();
                    filteredParts = allParts.filter(part => {
                        const searchString = [
                            part.input_mpn,
                            part.input_manufacturer,
                            part.found_mpn,
                            part.found_manufacturer,
                            ...(part.distributors || []).map(d => d.name)
                        ].join(' ').toLowerCase();
                        return searchString.includes(searchTerm);
                    });
                    currentPage = 1;
                    updateDisplay();
                });
                
                // Filter buttons
                document.querySelectorAll('.filter-btn').forEach(btn => {
                    btn.addEventListener('click', function() {
                        document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
                        this.classList.add('active');
                        
                        const filter = this.dataset.filter;
                        if (filter === 'all') {
                            filteredParts = [...allParts];
                        } else {
                            filteredParts = allParts.filter(part => part.match_type === filter);
                        }
                        currentPage = 1;
                        updateDisplay();
                    });
                });
            }
            
            function updateDisplay() {
                const container = document.getElementById('partsContainer');
                const start = (currentPage - 1) * partsPerPage;
                const end = start + partsPerPage;
                const partsToShow = filteredParts.slice(start, end);
                
                container.innerHTML = '';
                
                partsToShow.forEach((part, index) => {
                    const partCard = createPartCard(part, start + index);
                    container.appendChild(partCard);
                });
                
                // Update pagination
                document.getElementById('visibleCount').textContent = partsToShow.length;
                document.getElementById('totalCount').textContent = filteredParts.length;
                document.getElementById('currentPage').textContent = currentPage;
                document.getElementById('totalPages').textContent = Math.ceil(filteredParts.length / partsPerPage);
                
                document.getElementById('prevBtn').disabled = currentPage === 1;
                document.getElementById('nextBtn').disabled = currentPage >= Math.ceil(filteredParts.length / partsPerPage);
            }
            
            function createPartCard(part, index) {
                const card = document.createElement('div');
                card.className = 'part-card';
                card.dataset.index = index;
                
                const matchClass = 'match-' + part.match_type;
                
                let distributorHtml = '';
                if (part.distributors && part.distributors.length > 0) {
                    distributorHtml = '<div class="distributors-section"><h4>Distributors (' + part.distributors.length + ')</h4><div class="distributors-grid">';
                    
                    part.distributors.forEach(dist => {
                        let priceHtml = '';
                        const pricing = dist.pricing || dist.estimated_pricing || [];
                        
                        if (pricing.length > 0) {
                            priceHtml = '<table class="price-table"><thead><tr><th>Qty</th><th>Price</th></tr></thead><tbody>';
                            pricing.slice(0, 3).forEach(tier => {
                                if (tier.price !== undefined) {
                                    priceHtml += '<tr><td>' + tier.quantity.toLocaleString() + '</td><td>$' + tier.price.toFixed(4) + '</td></tr>';
                                }
                            });
                            if (pricing.length > 3) {
                                priceHtml += '<tr><td colspan="2">...and ' + (pricing.length - 3) + ' more tiers</td></tr>';
                            }
                            priceHtml += '</tbody></table>';
                        }
                        
                        distributorHtml += '<div class="distributor-card">' +
                            '<div class="distributor-name">' + dist.name + '</div>' +
                            '<div>Stock: ' + (dist.stock || 'N/A') + '</div>' +
                            '<div>Lead: ' + (dist.lead_time || 'N/A') + '</div>' +
                            priceHtml +
                            '</div>';
                    });
                    
                    distributorHtml += '</div></div>';
                }
                
                card.innerHTML = 
                    '<div class="part-header" onclick="toggleCard(' + index + ')">' +
                        '<div class="part-title">' + part.input_manufacturer + ' - ' + part.input_mpn + '</div>' +
                        '<div style="display: flex; align-items: center; gap: 1rem;">' +
                            '<span class="match-badge ' + matchClass + '">' + part.match_type + '</span>' +
                            '<span class="expand-icon">▼</span>' +
                        '</div>' +
                    '</div>' +
                    '<div class="part-content">' +
                        '<div class="part-details">' +
                            (part.found_mpn ? 
                                '<div><strong>Found:</strong> ' + part.found_manufacturer + ' - ' + part.found_mpn + '</div>' +
                                '<div><strong>Part ID:</strong> ' + (part.part_id || 'N/A') + '</div>'
                            : '') +
                        '</div>' +
                        distributorHtml +
                    '</div>';
                
                return card;
            }
            
            function toggleCard(index) {
                const card = document.querySelector('[data-index="' + index + '"]');
                card.classList.toggle('expanded');
            }
            
            function changePage(direction) {
                currentPage += direction;
                updateDisplay();
                window.scrollTo(0, 0);
            }
            
            function exportToCSV() {
                let csv = 'Manufacturer,MPN,Match Type,Found Manufacturer,Found MPN,Distributors,Best Price\\n';
                
                filteredParts.forEach(part => {
                    const bestPrice = getBestPrice(part);
                    csv += '"' + part.input_manufacturer + '","' + part.input_mpn + '","' + 
                        part.match_type + '","' + (part.found_manufacturer || '') + '","' + 
                        (part.found_mpn || '') + '","' + 
                        (part.distributors ? part.distributors.length : 0) + '","' + 
                        bestPrice + '"\\n';
                });
                
                const blob = new Blob([csv], { type: 'text/csv' });
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = 'pas_search_results.csv';
                a.click();
            }
            
            function getBestPrice(part) {
                if (!part.distributors) return 'N/A';
                
                let bestPrice = Infinity;
                part.distributors.forEach(dist => {
                    const pricing = dist.pricing || dist.estimated_pricing || [];
                    pricing.forEach(tier => {
                        if (tier.price && tier.price < bestPrice) {
                            bestPrice = tier.price;
                        }
                    });
                });
                
                return bestPrice === Infinity ? 'N/A' : '$' + bestPrice.toFixed(4);
            }
        </script>
    </body>
    </html>
        """
        
        # Save HTML file
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"Siemens-branded HTML report generated: {output_file}")
    
    def process_csv(self, input_file: str, output_file: str):
        """
        Process CSV file and create both Excel and HTML outputs with enriched results
        """
        # Clear raw responses for new processing
        self.raw_responses = []
        
        # Scan for pricing enrichers if enriching is enabled
        pricing_enrichers = {}
        if self.config.get('EnableEnriching', True):
            pricing_enrichers = self.scan_for_pricing_enrichers()
        
        # Detect encoding
        encoding = self._detect_encoding(input_file)
        
        # Read input CSV
        logger.info(f"\nReading input file: {input_file} with encoding: {encoding}")
        try:
            df = pd.read_csv(input_file, encoding=encoding)
        except Exception:
            # Try common encodings
            for enc in ['latin-1', 'iso-8859-1', 'cp1252', 'utf-16']:
                try:
                    df = pd.read_csv(input_file, encoding=enc)
                    logger.info(f"Successfully read with {enc} encoding")
                    break
                except:
                    continue
            else:
                raise ValueError(f"Could not read CSV file with any common encoding")
        
        # Standardize column names
        df.columns = df.columns.str.strip()
        
        # Find MPN and MFG columns
        mpn_col = 'MPN' if 'MPN' in df.columns else None
        mfr_col = 'MFG' if 'MFG' in df.columns else None
        
        if not mpn_col:
            for col in df.columns:
                if col.upper() == 'MPN':
                    mpn_col = col
                    break
        
        if not mfr_col:
            for col in df.columns:
                if col.upper() == 'MFG':
                    mfr_col = col
                    break
        
        if not mpn_col or not mfr_col:
            raise ValueError(f"Input CSV must have 'MPN' and 'MFG' columns. Found: {list(df.columns)}")
        
        logger.info(f"Using columns: MPN='{mpn_col}', MFG='{mfr_col}'")
        
        # Prepare results for Excel
        excel_results = []
        # Prepare results for HTML
        html_results = []
        
        total_parts = len(df)
        logger.info(f"\nProcessing {total_parts} parts...")
        
        for idx, row in df.iterrows():
            mpn = str(row[mpn_col]).strip() if pd.notna(row[mpn_col]) else ''
            mfr = str(row[mfr_col]).strip() if pd.notna(row[mfr_col]) else ''
            
            if not mpn or not mfr:
                logger.warning(f"Skipping row {idx+1}: Missing MPN or Manufacturer")
                continue
            
            logger.info(f"\n{'='*60}")
            logger.info(f"Searching [{idx+1}/{total_parts}]: {mfr} {mpn}")
            logger.info(f"{'='*60}")
            
            # Use search with enriching
            search_result, match_type = self.search_part_with_enriching(mpn, mfr, pricing_enrichers)
            
            # Store raw response if enabled
            if self.config.get('EnableRawOutput', False):
                self.raw_responses.append({
                    'input_mpn': mpn,
                    'input_manufacturer': mfr,
                    'match_type': match_type,
                    'timestamp': datetime.now().isoformat(),
                    'response': search_result
                })
            
            # Prepare HTML result
            html_result = {
                'input_mpn': mpn,
                'input_manufacturer': mfr,
                'match_type': match_type,
                'distributors': []
            }
            
            if 'error' in search_result:
                excel_results.append({
                    'Input_MPN': mpn,
                    'Input_Manufacturer': mfr,
                    'Match_Type': 'error',
                    'Found_MPN': '',
                    'Found_Manufacturer': '',
                    'Part_ID': '',
                    'Description': f"Error: {search_result['error']}",
                    'Total_Results': 0
                })
                html_results.append(html_result)
            elif match_type != 'no_match' and search_result.get('result'):
                parts = search_result['result'].get('results', [])
                total_count = search_result['result'].get('totalCount', len(parts))
                
                if not parts:
                    excel_results.append({
                        'Input_MPN': mpn,
                        'Input_Manufacturer': mfr,
                        'Match_Type': 'no_match',
                        'Found_MPN': '',
                        'Found_Manufacturer': '',
                        'Part_ID': '',
                        'Description': 'No matching parts found',
                        'Total_Results': 0
                    })
                    html_results.append(html_result)
                else:
                    # Use only the first match for HTML
                    first_part = parts[0]
                    part = first_part.get('searchProviderPart', {})
                    
                    html_result['found_mpn'] = part.get('manufacturerPartNumber', '')
                    html_result['found_manufacturer'] = part.get('manufacturerName', '')
                    html_result['part_id'] = part.get('partId', '')
                    
                    # Extract enriching data for HTML
                    enriched = self.extract_enriching_data_improved(first_part)
                    html_result['distributors'] = enriched.get('distributors', [])
                    
                    html_results.append(html_result)
                    
                    # Add to Excel results
                    excel_results.append({
                        'Input_MPN': mpn,
                        'Input_Manufacturer': mfr,
                        'Match_Type': match_type,
                        'Found_MPN': part.get('manufacturerPartNumber', ''),
                        'Found_Manufacturer': part.get('manufacturerName', ''),
                        'Part_ID': part.get('partId', ''),
                        'Description': f"Found {total_count} matches",
                        'Total_Results': total_count
                    })
            else:
                excel_results.append({
                    'Input_MPN': mpn,
                    'Input_Manufacturer': mfr,
                    'Match_Type': 'no_match',
                    'Found_MPN': '',
                    'Found_Manufacturer': '',
                    'Part_ID': '',
                    'Description': 'No matching parts found',
                    'Total_Results': 0
                })
                html_results.append(html_result)
        
        # Create DataFrame and export to Excel
        results_df = pd.DataFrame(excel_results)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Search Results', index=False)
            
            # Format the Excel file
            workbook = writer.book
            worksheet = writer.sheets['Search Results']
            
            # Add colors for match types
            for row in range(2, len(results_df) + 2):
                match_type = worksheet.cell(row=row, column=3).value
                if match_type == 'exact':
                    fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif match_type == 'partial':
                    fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif match_type in ['no_match', 'error']:
                    fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    continue
                worksheet.cell(row=row, column=3).fill = fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"\nExcel results exported to: {output_file}")
        
        # Generate HTML report
        html_file = output_file.replace('.xlsx', '_report.html')
        self.generate_html_report(html_results, html_file)
        
        # Save raw responses if enabled
        if self.config.get('EnableRawOutput', False) and self.raw_responses:
            raw_output_file = output_file.replace('.xlsx', '_raw.json')
            with open(raw_output_file, 'w') as f:
                json.dump(self.raw_responses, f, indent=2)
            logger.info(f"Raw API responses saved to: {raw_output_file}")
        
        # Print summary
        summary = results_df['Match_Type'].value_counts()
        logger.info("\n=== SUMMARY ===")
        logger.info(f"Total parts searched: {total_parts}")
        for match_type, count in summary.items():
            logger.info(f"{match_type}: {count}")


def main():
    """
    Main function with multiple usage options
    """
    print("\n" + "="*60)
    print("PAS API Part Search Tool with Enriching and HTML Report")
    print("="*60 + "\n")
    
    # Option to use a pre-obtained bearer token
    use_token = input("Do you want to provide a bearer token directly? (y/n, default: n): ").strip().lower()
    
    if use_token == 'y':
        bearer_token = input("Enter your bearer token: ").strip()
        if not bearer_token:
            logger.error("Bearer token cannot be empty")
            return
        client = PASAPIClient(bearer_token=bearer_token)
        logger.info("Using provided bearer token")
    else:
        client = PASAPIClient()
        logger.info("Client will obtain its own token using credentials")
    
    # Ask about enriching
    enable_enriching = input("Enable enriching data (lifecycle, pricing, etc.)? (y/n, default: y): ").strip().lower()
    if enable_enriching != 'n':
        client.config['EnableEnriching'] = True
        logger.info("Enriching enabled - will scan for all available enrichers including pricing")
    else:
        client.config['EnableEnriching'] = False
        logger.info("Enriching disabled - basic search only")
    
    # Process CSV file
    input_file = input("Enter input CSV filename (default: input.csv): ").strip() or "input.csv"
    output_file = input("Enter output Excel filename (default: search_results.xlsx): ").strip() or "search_results.xlsx"
    
    if not os.path.exists(input_file):
        logger.error(f"Input file '{input_file}' not found!")
        logger.info("Creating a sample input file...")
        sample_data = {
            'MFG': ['ROHM', 'Texas Instruments', 'STMicroelectronics'],
            'MPN': ['UDZVTE-176.2B', 'LM358DR', 'STM32F103C8T6']
        }
        pd.DataFrame(sample_data).to_csv(input_file, index=False)
        logger.info(f"Sample file '{input_file}' created. Please edit it with your parts and run again.")
        return
    
    try:
        client.process_csv(input_file, output_file)
        logger.info("\nProcessing complete!")
        logger.info(f"Check '{output_file}' for basic results")
        logger.info(f"Check '{output_file.replace('.xlsx', '_report.html')}' for detailed HTML report with pricing")
        if client.config.get('EnableRawOutput', False):
            logger.info(f"Check '{output_file.replace('.xlsx', '_raw.json')}' for raw API responses")
    except Exception as e:
        logger.error(f"Processing failed: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()